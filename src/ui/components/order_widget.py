"""
Order widget component for displaying and managing orders.
"""
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                           QPushButton, QScrollArea, QFrame, QDialog,
                           QLineEdit, QTextEdit, QMessageBox, QComboBox,
                           QTableWidget, QTableWidgetItem, QHeaderView,
                           QTabWidget, QSplitter, QGroupBox, QGridLayout,
                           QDateEdit, QProgressBar, QApplication)
from PyQt5.QtCore import Qt, pyqtSignal, QTimer, QDate
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtWidgets import QApplication
from datetime import datetime, date
from controllers.order_controller import OrderController
from models.order import OrderStatus
from models.user import User
import logging

# Excel imports (optional - will be imported when needed)
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    # Create dummy classes to avoid import errors
    class Font:
        def __init__(self, **kwargs):
            pass
    class PatternFill:
        def __init__(self, **kwargs):
            pass
    class Alignment:
        def __init__(self, **kwargs):
            pass
    class Border:
        def __init__(self, **kwargs):
            pass
    class Side:
        def __init__(self, **kwargs):
            pass

logger = logging.getLogger(__name__)

class OrderCard(QFrame):
    """Widget for displaying an individual order."""
    
    # Signals
    order_completed = pyqtSignal(object)  # Emitted when order is completed
    order_cancelled = pyqtSignal(object)  # Emitted when order is cancelled
    order_edited = pyqtSignal(object)     # Emitted when order is edited
    order_clicked = pyqtSignal(object)    # Emitted when order card is clicked
    
    def __init__(self, order, user: User, parent=None):
        super().__init__(parent)
        self.order = order
        self.user = user
        self.order_controller = OrderController()
        self.init_ui()
        
    def init_ui(self):
        """Initialize the user interface."""
        self.setFrameStyle(QFrame.StyledPanel)
        
        # Enable mouse tracking for hover effects
        self.setMouseTracking(True)
        
        # Set cursor and styling based on order status
        if self.order.status == OrderStatus.ACTIVE:
            self.setCursor(Qt.PointingHandCursor)  # Show pointer cursor for active orders
            # Active orders get full hover effect
            self.setStyleSheet('''
                OrderCard {
                    background-color: #ffffff;
                    border: 2px solid #e0e0e0;
                    border-radius: 12px;
                    padding: 15px;
                    margin: 5px;
                    min-width: 320px;
                    max-width: 380px;
                    min-height: 200px;
                }
                OrderCard:hover {
                    border-color: #3498db;
                    background-color: #f8f9fa;
                }
            ''')
        elif self.order.status == OrderStatus.COMPLETED:
            self.setCursor(Qt.ArrowCursor)  # Show normal cursor for completed orders
            # Completed orders get minimal hover effect
            self.setStyleSheet('''
                OrderCard {
                    background-color: #f8f9fa;
                    border: 2px solid #bdc3c7;
                    border-radius: 12px;
                    padding: 15px;
                    margin: 5px;
                    min-width: 320px;
                    max-width: 380px;
                    min-height: 200px;
                }
                OrderCard:hover {
                    border-color: #95a5a6;
                    background-color: #ecf0f1;
                }
            ''')
        else:
            self.setCursor(Qt.ArrowCursor)  # Default cursor for other statuses
            # Default styling for other statuses
            self.setStyleSheet('''
                OrderCard {
                    background-color: #ffffff;
                    border: 2px solid #e0e0e0;
                    border-radius: 12px;
                    padding: 15px;
                    margin: 5px;
                    min-width: 320px;
                    max-width: 380px;
                    min-height: 200px;
                }
                OrderCard:hover {
                    border-color: #3498db;
                    background-color: #f8f9fa;
                }
            ''')
        
        # Add tooltip for completed orders
        if self.order.status == OrderStatus.COMPLETED:
            self.setToolTip("Completed orders cannot be loaded into cart")
        elif self.order.status == OrderStatus.ACTIVE:
            self.setToolTip("Click to load this active order into cart for modification or checkout")
        else:
            self.setToolTip("Click to load this order into cart for checkout")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)
        
        # Header with status only
        header_layout = QHBoxLayout()
        
        header_layout.addStretch()
        
        # Status badge
        status_label = QLabel(self.order.get_status_display())
        status_label.setStyleSheet(self._get_status_style())
        status_label.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(status_label)
        
        layout.addLayout(header_layout)
        
        # Customer name
        if self.order.customer_name:
            customer_label = QLabel(f"Customer: {self.order.customer_name}")
            customer_label.setStyleSheet("font-weight: bold; color: #34495e; margin: 5px 0; font-size: 13px;")
            layout.addWidget(customer_label)
        
        # Order items
        items_label = QLabel("Items:")
        items_label.setStyleSheet("font-weight: bold; margin-top: 10px; font-size: 13px;")
        layout.addWidget(items_label)
        
        items_text = ""
        order_items = self.order.get_order_items()
        
        # Get a session to access product categories
        from database.db_config import get_fresh_session
        session = get_fresh_session()
        
        try:
            for item in order_items:
                product = item['product']
                item_total = item['price'] * item['quantity']
                tax_rate = 0.0
                
                # Get product with category from session to avoid lazy loading issues
                try:
                    from models.product import Product
                    session_product = session.query(Product).filter_by(id=product.id).first()
                    if session_product and hasattr(session_product, 'category') and session_product.category:
                        tax_rate = getattr(session_product.category, 'tax_rate', 0.0)
                except Exception as e:
                    # If we can't get the category, just use 0 tax rate
                    tax_rate = 0.0
                
                items_text += f"• {product.name} x{item['quantity']} - ${item['price']:.2f}\n"
                items_text += f"  Subtotal: ${item_total:.2f}"
                if tax_rate > 0:
                    item_tax = item_total * (tax_rate / 100.0)
                    items_text += f" (Tax: {tax_rate}% = ${item_tax:.2f})"
                items_text += "\n"
                
                if item.get('notes'):
                    items_text += f"  Note: {item['notes']}\n"
        finally:
            session.close()
        
        if items_text:
            items_display = QLabel(items_text)
            items_display.setStyleSheet("color: #555; font-size: 12px; margin-left: 10px;")
            items_display.setWordWrap(True)
            layout.addWidget(items_display)
        
        # Order totals breakdown
        totals_layout = QVBoxLayout()
        totals_layout.setSpacing(2)
        
        # Subtotal
        subtotal_label = QLabel(f"Subtotal: ${self.order.subtotal:.2f}")
        subtotal_label.setStyleSheet("color: #555; font-size: 12px;")
        totals_layout.addWidget(subtotal_label)
        
        # Tax amount
        if self.order.tax_amount > 0:
            tax_label = QLabel(f"Tax: ${self.order.tax_amount:.2f}")
            tax_label.setStyleSheet("color: #e67e22; font-size: 12px;")
            totals_layout.addWidget(tax_label)
        
        # Discount amount
        if self.order.discount_amount > 0:
            discount_label = QLabel(f"Discount: -${self.order.discount_amount:.2f}")
            discount_label.setStyleSheet("color: #e74c3c; font-size: 12px;")
            totals_layout.addWidget(discount_label)
        
        # Total amount
        total_label = QLabel(f"Total: ${self.order.total_amount:.2f}")
        total_label.setStyleSheet("font-weight: bold; color: #27ae60; font-size: 15px; margin-top: 5px;")
        totals_layout.addWidget(total_label)
        
        layout.addLayout(totals_layout)
        
        # Time created
        time_label = QLabel(f"Created: {self.order.created_at.strftime('%I:%M:%S %p')}")
        time_label.setStyleSheet("color: #7f8c8d; font-size: 11px;")
        layout.addWidget(time_label)
        
        # Notes if any
        if self.order.notes:
            notes_label = QLabel(f"Notes: {self.order.notes}")
            notes_label.setStyleSheet("color: #e67e22; font-style: italic; font-size: 11px; margin-top: 5px;")
            notes_label.setWordWrap(True)
            layout.addWidget(notes_label)
        
        # Action buttons (only for active orders)
        if self.order.status == OrderStatus.ACTIVE:
            btn_layout = QHBoxLayout()
            btn_layout.setSpacing(8)
            
            edit_btn = QPushButton("Edit")
            edit_btn.setStyleSheet("""
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 12px;
                font-weight: bold;
                min-height: 30px;
            """)
            edit_btn.clicked.connect(self.edit_order)
            btn_layout.addWidget(edit_btn)
            
            cancel_btn = QPushButton("Cancel")
            cancel_btn.setStyleSheet("""
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 12px;
                font-weight: bold;
                min-height: 30px;
            """)
            cancel_btn.clicked.connect(self.cancel_order)
            btn_layout.addWidget(cancel_btn)
            
            layout.addLayout(btn_layout)
            
            # Add a note for active orders that they can be clicked to load into cart
            note_label = QLabel("💡 Click anywhere on this card to load into cart for modification or checkout")
            note_label.setStyleSheet("""
                color: #3498db;
                font-style: italic;
                font-size: 11px;
                margin-top: 5px;
                text-align: center;
            """)
            layout.addWidget(note_label)
        elif self.order.status == OrderStatus.COMPLETED:
            # Add a note for completed orders that they can be clicked to load into cart
            note_label = QLabel("💡 Click to load into cart for checkout")
            note_label.setStyleSheet("""
                color: #f39c12;
                font-style: italic;
                font-size: 11px;
                margin-top: 5px;
                text-align: center;
            """)
            layout.addWidget(note_label)
    
    def _get_status_style(self):
        """Get CSS style for status badge."""
        status_styles = {
            OrderStatus.ACTIVE: "background-color: #3498db; color: white; padding: 3px 8px; border-radius: 10px; font-size: 10px; font-weight: bold;",
            OrderStatus.COMPLETED: "background-color: #27ae60; color: white; padding: 3px 8px; border-radius: 10px; font-size: 10px; font-weight: bold;",
            OrderStatus.CANCELLED: "background-color: #e74c3c; color: white; padding: 3px 8px; border-radius: 10px; font-size: 10px; font-weight: bold;",
            OrderStatus.PENDING: "background-color: #f39c12; color: white; padding: 3px 8px; border-radius: 10px; font-size: 10px; font-weight: bold;"
        }
        return status_styles.get(self.order.status, "")
    
    def edit_order(self):
        """Edit the order."""
        dialog = OrderEditDialog(self.order, self.user, self)
        if dialog.exec_() == QDialog.Accepted:
            self.order_edited.emit(self.order)
    
    def complete_order(self):
        """Complete the order."""
        reply = QMessageBox.question(
            self, 
            "Complete Order", 
            f"Are you sure you want to complete order #{self.order.order_number}?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if self.order_controller.complete_order(self.order):
                self.order_completed.emit(self.order)
                QMessageBox.information(self, "Success", "Order completed successfully!")
            else:
                QMessageBox.warning(self, "Error", "Failed to complete order.")
    
    def cancel_order(self):
        """Cancel the order."""
        dialog = OrderCancelDialog(self.order, self.user, self)
        if dialog.exec_() == QDialog.Accepted:
            if self.order_controller.cancel_order(self.order, self.user, dialog.reason):
                self.order_cancelled.emit(self.order)
                QMessageBox.information(self, "Success", "Order cancelled successfully!")
            else:
                QMessageBox.warning(self, "Error", "Failed to cancel order.")
    
    def mousePressEvent(self, event):
        """Handle mouse click events."""
        if event.button() == Qt.LeftButton:
            # Only allow clicking on active orders to load into cart
            # Completed orders should not be editable or loadable
            if self.order.status == OrderStatus.ACTIVE:
                self.order_clicked.emit(self.order)
            elif self.order.status == OrderStatus.COMPLETED:
                # Show message that completed orders cannot be loaded
                QMessageBox.information(
                    self, 
                    "Order Status", 
                    f"Order {self.order.order_number} is already completed and cannot be loaded into cart.\n\n"
                    "Completed orders are final and cannot be modified.\n\n"
                    "To process a completed order for payment, please create a new sale."
                )
            else:
                # For other statuses (cancelled, pending), show appropriate message
                QMessageBox.information(
                    self,
                    "Order Status",
                    f"Order {self.order.order_number} has status '{self.order.get_status_display()}' and cannot be loaded into cart.\n\n"
                    "Only active orders can be loaded for modification or checkout."
                )
        try:
            super().mousePressEvent(event)
        except RuntimeError:
            # Widget has been deleted, ignore
            pass


class OrderEditDialog(QDialog):
    """Dialog for editing order details."""
    
    def __init__(self, order, user: User, parent=None):
        super().__init__(parent)
        self.order = order
        self.user = user
        self.order_controller = OrderController()
        self.setWindowTitle("Edit Order")
        self.setFixedSize(400, 300)
        self.init_ui()
    
    def init_ui(self):
        """Initialize the user interface."""
        layout = QVBoxLayout(self)
        
        # Customer name
        layout.addWidget(QLabel("Customer Name:"))
        self.customer_name_input = QLineEdit(self.order.customer_name or "")
        layout.addWidget(self.customer_name_input)
        
        # Notes
        layout.addWidget(QLabel("Notes:"))
        self.notes_input = QTextEdit()
        self.notes_input.setPlainText(self.order.notes or "")
        self.notes_input.setMaximumHeight(100)
        layout.addWidget(self.notes_input)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        save_btn = QPushButton("Save")
        save_btn.clicked.connect(self.save_changes)
        save_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px 16px; border-radius: 5px;")
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("background-color: #95a5a6; color: white; padding: 8px 16px; border-radius: 5px;")
        
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
    
    def save_changes(self):
        """Save the changes to the order."""
        customer_name = self.customer_name_input.text().strip()
        notes = self.notes_input.toPlainText().strip()
        
        success = True
        if customer_name != (self.order.customer_name or ""):
            success &= self.order_controller.update_order_customer_name(self.order, customer_name)
        
        if notes != (self.order.notes or ""):
            success &= self.order_controller.update_order_notes(self.order, notes)
        
        if success:
            self.accept()
        else:
            QMessageBox.warning(self, "Error", "Failed to save changes.")


class OrderCancelDialog(QDialog):
    """Dialog for cancelling an order."""
    
    def __init__(self, order, user: User, parent=None):
        super().__init__(parent)
        self.order = order
        self.user = user
        self.reason = ""
        self.setWindowTitle("Cancel Order")
        self.setFixedSize(400, 200)
        self.init_ui()
    
    def init_ui(self):
        """Initialize the user interface."""
        layout = QVBoxLayout(self)
        
        layout.addWidget(QLabel("Are you sure you want to cancel this order?"))
        
        # Reason for cancellation
        layout.addWidget(QLabel("Reason for cancellation (optional):"))
        self.reason_input = QTextEdit()
        self.reason_input.setMaximumHeight(80)
        layout.addWidget(self.reason_input)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        confirm_btn = QPushButton("Confirm Cancel")
        confirm_btn.clicked.connect(self.confirm_cancel)
        confirm_btn.setStyleSheet("background-color: #e74c3c; color: white; padding: 8px 16px; border-radius: 5px;")
        
        cancel_btn = QPushButton("Keep Order")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("background-color: #95a5a6; color: white; padding: 8px 16px; border-radius: 5px;")
        
        btn_layout.addWidget(confirm_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
    
    def confirm_cancel(self):
        """Confirm the cancellation."""
        self.reason = self.reason_input.toPlainText().strip()
        self.accept()


class OrderResetDialog(QDialog):
    """Dialog for resetting the order manager."""
    
    def __init__(self, user: User, parent=None):
        super().__init__(parent)
        self.user = user
        self.reset_type = "all"
        self.setWindowTitle("Reset Order Manager")
        self.setFixedSize(450, 350)
        self.init_ui()
    
    def init_ui(self):
        """Initialize the user interface."""
        layout = QVBoxLayout(self)
        
        # Warning message
        warning_label = QLabel("⚠️ WARNING: This action cannot be undone!")
        warning_label.setStyleSheet("""
            color: #e74c3c;
            font-weight: bold;
            font-size: 14px;
            padding: 10px;
            background-color: #fdf2f2;
            border: 1px solid #e74c3c;
            border-radius: 5px;
        """)
        layout.addWidget(warning_label)
        
        # Archive info
        archive_info = QLabel("📋 Note: Orders will be archived instead of deleted.")
        archive_info.setStyleSheet("""
            color: #27ae60;
            font-weight: bold;
            font-size: 12px;
            padding: 8px;
            background-color: #e8f5e8;
            border: 1px solid #27ae60;
            border-radius: 5px;
            margin: 5px 0;
        """)
        layout.addWidget(archive_info)
        
        # Description
        desc_label = QLabel("Select what you want to reset in the order manager:")
        desc_label.setStyleSheet("font-size: 12px; color: #666; margin: 10px 0;")
        layout.addWidget(desc_label)
        
        # Reset options
        from PyQt5.QtWidgets import QRadioButton, QButtonGroup
        
        self.radio_group = QButtonGroup()
        
        # All orders option
        all_radio = QRadioButton("Archive All Orders (Active, Completed, Cancelled, Old)")
        all_radio.setChecked(True)
        all_radio.setStyleSheet("font-weight: bold; color: #e74c3c;")
        self.radio_group.addButton(all_radio, 0)
        layout.addWidget(all_radio)
        
        # Active orders option
        active_radio = QRadioButton("Archive Active Orders Only")
        active_radio.setStyleSheet("color: #3498db;")
        self.radio_group.addButton(active_radio, 1)
        layout.addWidget(active_radio)
        
        # Completed orders option
        completed_radio = QRadioButton("Archive Completed Orders Only")
        completed_radio.setStyleSheet("color: #27ae60;")
        self.radio_group.addButton(completed_radio, 2)
        layout.addWidget(completed_radio)
        
        # Cancelled orders option
        cancelled_radio = QRadioButton("Archive Cancelled Orders Only")
        cancelled_radio.setStyleSheet("color: #f39c12;")
        self.radio_group.addButton(cancelled_radio, 3)
        layout.addWidget(cancelled_radio)
        
        # Old orders option
        old_radio = QRadioButton("Clean Up Old Orders (Older than 24 hours)")
        old_radio.setStyleSheet("color: #9b59b6;")
        self.radio_group.addButton(old_radio, 4)
        layout.addWidget(old_radio)
        
        # Connect radio buttons
        self.radio_group.buttonClicked.connect(self.on_radio_changed)
        
        # Current stats
        stats_group = QGroupBox("Current Order Statistics")
        stats_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ddd;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
        """)
        stats_layout = QVBoxLayout(stats_group)
        
        self.stats_label = QLabel("Loading statistics...")
        self.stats_label.setStyleSheet("font-size: 11px; color: #666;")
        stats_layout.addWidget(self.stats_label)
        
        layout.addWidget(stats_group)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        reset_btn = QPushButton("Reset Orders")
        reset_btn.clicked.connect(self.confirm_reset)
        reset_btn.setStyleSheet("""
            background-color: #e74c3c;
            color: white;
            padding: 10px 20px;
            border-radius: 5px;
            font-weight: bold;
        """)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            background-color: #95a5a6;
            color: white;
            padding: 10px 20px;
            border-radius: 5px;
        """)
        
        btn_layout.addWidget(reset_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        
        # Load current stats
        self.load_stats()
    
    def on_radio_changed(self, button):
        """Handle radio button selection change."""
        button_id = self.radio_group.id(button)
        reset_types = ["all", "active", "completed", "cancelled", "old"]
        self.reset_type = reset_types[button_id]
    
    def load_stats(self):
        """Load and display current order statistics."""
        try:
            from controllers.order_controller import OrderController
            order_controller = OrderController()
            stats = order_controller.get_order_manager_stats()
            
            stats_text = f"""
            Active Orders: {stats['active_orders']}
            Completed Orders: {stats['completed_orders']}
            Cancelled Orders: {stats['cancelled_orders']}
            Today's Orders: {stats['today_orders']}
            Total Orders: {stats['total_orders']}
            """
            
            self.stats_label.setText(stats_text)
        except Exception as e:
            self.stats_label.setText("Error loading statistics")
    
    def confirm_reset(self):
        """Confirm the reset operation."""
        # Show confirmation dialog
        reset_types = {
            "all": "ALL ORDERS (Active, Completed, Cancelled, and Old)",
            "active": "ACTIVE ORDERS ONLY",
            "completed": "COMPLETED ORDERS ONLY", 
            "cancelled": "CANCELLED ORDERS ONLY",
            "old": "OLD ORDERS (Older than 24 hours)"
        }
        
        reset_type_name = reset_types.get(self.reset_type, self.reset_type.upper())
        
        reply = QMessageBox.question(
            self,
            "Confirm Reset",
            f"Are you absolutely sure you want to archive {reset_type_name}?\n\n"
            "Orders will be archived (not deleted).\n"
            "This action cannot be undone!",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.accept()
        else:
            self.reject()


class OrderManagementWidget(QWidget):
    """Main widget for managing orders with tabs for different statuses."""
    
    # Signals
    order_updated = pyqtSignal()  # Emitted when any order is updated
    
    def __init__(self, user: User, parent=None):
        super().__init__(parent)
        self.user = user
        self.order_controller = OrderController()
        self.order_cards = {}  # Store order cards by order ID
        self.init_ui()
        self.refresh_orders()
        
        # Auto-refresh timer
        self.refresh_timer = QTimer()
        self.refresh_timer.timeout.connect(self.refresh_orders)
        self.refresh_timer.start(30000)  # Refresh every 30 seconds
    
    def init_ui(self):
        """Initialize the user interface."""
        layout = QVBoxLayout(self)
        
        # Header
        header_layout = QHBoxLayout()
        title_label = QLabel("Order Management")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setStyleSheet("color: #2c3e50; margin-bottom: 10px;")
        header_layout.addWidget(title_label)
        
        # Today's order count
        self.today_count_label = QLabel("Today: 0")
        self.today_count_label.setStyleSheet("""
            color: #27ae60;
            font-weight: bold;
            font-size: 14px;
            padding: 5px 10px;
            background-color: #e8f5e8;
            border-radius: 5px;
        """)
        header_layout.addWidget(self.today_count_label)
        
        header_layout.addStretch()
        
        # Refresh button
        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.refresh_orders)
        refresh_btn.setStyleSheet("""
            background-color: #3498db;
            color: white;
            border: none;
            border-radius: 5px;
            padding: 8px 16px;
            font-weight: bold;
        """)
        header_layout.addWidget(refresh_btn)
        

        
        # Day Report button
        day_report_btn = QPushButton("📊 Day Report")
        day_report_btn.clicked.connect(self.show_day_report_dialog)
        day_report_btn.setStyleSheet("""
            background-color: #9b59b6;
            color: white;
            border: none;
            border-radius: 5px;
            padding: 8px 16px;
            font-weight: bold;
        """)
        header_layout.addWidget(day_report_btn)
        
        # Reset button - only show for admin users
        if self.user.role.value == "admin":
            reset_btn = QPushButton("Reset")
            reset_btn.clicked.connect(self.show_reset_dialog)
            reset_btn.setStyleSheet("""
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: bold;
            """)
            header_layout.addWidget(reset_btn)
        
        layout.addLayout(header_layout)
        
        # Tab widget for different order statuses
        self.tab_widget = QTabWidget()
        
        # Active orders tab
        self.active_tab = self.create_orders_tab("Active Orders")
        self.tab_widget.addTab(self.active_tab, "Active Orders")
        
        # Completed orders tab
        self.completed_tab = self.create_orders_tab("Completed Orders")
        self.tab_widget.addTab(self.completed_tab, "Completed Orders")
        
        # Cancelled orders tab
        self.cancelled_tab = self.create_orders_tab("Cancelled Orders")
        self.tab_widget.addTab(self.cancelled_tab, "Cancelled Orders")
        
        layout.addWidget(self.tab_widget)
    
    def create_orders_tab(self, title):
        """Create a tab for displaying orders of a specific status."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        # Scroll area for order cards
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        
        # Container for order cards with grid layout
        orders_container = QWidget()
        orders_layout = QGridLayout(orders_container)
        orders_layout.setSpacing(15)
        orders_layout.setContentsMargins(10, 10, 10, 10)
        
        scroll.setWidget(orders_container)
        layout.addWidget(scroll)
        
        # Store reference to the container for later use
        tab.orders_container = orders_container
        tab.orders_layout = orders_layout
        tab.title = title  # Store title for reference
        
        return tab
    
    def refresh_orders(self):
        """Refresh all orders."""
        # Clear existing order cards
        for card in self.order_cards.values():
            card.setParent(None)
            card.deleteLater()
        self.order_cards.clear()
        
        # Get orders by status
        active_orders = self.order_controller.get_active_orders()
        completed_orders = self.order_controller.get_completed_orders()
        cancelled_orders = self.order_controller.get_cancelled_orders()
        
        # Add order cards to appropriate tabs
        self.add_order_cards_to_tab(self.active_tab, active_orders)
        self.add_order_cards_to_tab(self.completed_tab, completed_orders)
        self.add_order_cards_to_tab(self.cancelled_tab, cancelled_orders)
        
        # Update today's order count
        self.update_today_count()
        
        # Emit signal that orders were updated
        self.order_updated.emit()
    
    def update_today_count(self):
        """Update the today's order count display."""
        try:
            today_count = self.order_controller.get_today_order_count()
            self.today_count_label.setText(f"Today: {today_count}")
        except Exception as e:
            logger.error(f"Error updating today's order count: {e}")
            self.today_count_label.setText("Today: ?")
    
    def force_refresh(self):
        """Force a complete refresh of the order list."""
        self.refresh_orders()
    
    def add_order_cards_to_tab(self, tab, orders):
        """Add order cards to a specific tab using responsive grid layout."""
        # Clear existing widgets in the tab
        while tab.orders_layout.count():
            item = tab.orders_layout.takeAt(0)
            if item.widget():
                item.widget().setParent(None)
        
        if not orders:
            # Show "No orders" message
            title = getattr(tab, 'title', 'orders').lower()
            no_orders_label = QLabel(f"No {title} found")
            no_orders_label.setAlignment(Qt.AlignCenter)
            no_orders_label.setStyleSheet("font-size: 16px; color: #666; padding: 40px; background-color: #f8f9fa; border-radius: 10px;")
            tab.orders_layout.addWidget(no_orders_label, 0, 0, 1, 3)  # Span multiple columns
            return
        
        # Calculate responsive grid columns
        available_width = self.width() - 50  # Account for margins
        card_width = 350  # Approximate card width
        max_cols = max(1, min(4, int(available_width / card_width)))  # Limit to 4 columns max
        
        row = 0
        col = 0
        
        # Add new order cards
        for order in orders:
            card = OrderCard(order, self.user)
            card.order_completed.connect(self.on_order_completed)
            card.order_cancelled.connect(self.on_order_cancelled)
            card.order_edited.connect(self.on_order_edited)
            card.order_clicked.connect(self.on_order_clicked)
            
            tab.orders_layout.addWidget(card, row, col)
            self.order_cards[order.id] = card
            
            col += 1
            if col >= max_cols:
                col = 0
                row += 1
        
        # Add stretch to fill remaining space
        tab.orders_layout.setRowStretch(row + 1, 1)
        for i in range(max_cols):
            tab.orders_layout.setColumnStretch(i, 1)
    
    def on_order_completed(self, order):
        """Handle order completion."""
        self.refresh_orders()
        self.order_updated.emit()
    
    def on_order_cancelled(self, order):
        """Handle order cancellation."""
        self.refresh_orders()
        self.order_updated.emit()
    
    def on_order_edited(self, order):
        """Handle order editing."""
        self.refresh_orders()
        self.order_updated.emit()
    
    def on_order_clicked(self, order):
        """Handle order click - load order into cart."""
        try:
            # Find the main window to access the cart
            parent = self.parent()
            cart_widget = None
            
            # Try different ways to find the cart widget
            while parent:
                # Method 1: Check if this is the main window with pos_widget
                if hasattr(parent, 'pos_widget') and hasattr(parent.pos_widget, 'cart_widget'):
                    cart_widget = parent.pos_widget.cart_widget
                    break
                
                # Method 2: Check if this is the stacked widget parent
                if hasattr(parent, 'parent') and parent.parent():
                    main_parent = parent.parent()
                    if hasattr(main_parent, 'pos_widget') and hasattr(main_parent.pos_widget, 'cart_widget'):
                        cart_widget = main_parent.pos_widget.cart_widget
                        break
                
                # Method 3: Check if this is the main window directly
                if hasattr(parent, 'cart_widget'):
                    cart_widget = parent.cart_widget
                    break
                
                # Method 4: Look for the main window class
                if hasattr(parent, '__class__') and 'MainWindow' in parent.__class__.__name__:
                    if hasattr(parent, 'pos_widget') and hasattr(parent.pos_widget, 'cart_widget'):
                        cart_widget = parent.pos_widget.cart_widget
                        break
                
                parent = parent.parent()
            
            if cart_widget and hasattr(cart_widget, 'load_order'):
                # Load the order into the cart
                cart_widget.load_order(order)
                
                # Switch to POS tab if we can find the sidebar
                parent = self.parent()
                while parent:
                    if hasattr(parent, 'sidebar'):
                        parent.sidebar.setCurrentRow(0)  # Switch to POS tab
                        break
                    parent = parent.parent()
                
                # Show appropriate message based on order status
                if order.status.value == "completed":
                    QMessageBox.information(
                        self, 
                        "Order Loaded", 
                        f"Completed order {order.order_number} loaded into cart!\n\n"
                        f"Customer: {order.customer_name or 'No name'}\n"
                        f"Total: ${order.total_amount:.2f}\n\n"
                        f"Click 'Checkout' to process the sale."
                    )
                else:
                    QMessageBox.information(
                        self, 
                        "Order Loaded", 
                        f"Order {order.order_number} loaded into cart!\n\n"
                        f"Customer: {order.customer_name or 'No name'}\n"
                        f"Status: {order.get_status_display()}\n\n"
                        f"You can now modify the order or proceed to checkout."
                    )
            else:
                QMessageBox.warning(
                    self, 
                    "Error", 
                    "Could not find cart widget or cart widget does not support loading orders.\n\n"
                    "Please ensure you are using the enhanced cart widget."
                )
                
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Error", 
                f"An error occurred while loading the order into cart:\n{str(e)}"
            )
    


    def show_day_report_dialog(self):
        """Show the day report dialog."""
        dialog = DayReportDialog(self.user, self)
        dialog.exec_()
    
    def show_reset_dialog(self):
        """Show the order reset dialog."""
        dialog = OrderResetDialog(self.user, self)
        if dialog.exec_() == QDialog.Accepted:
            reset_type = dialog.reset_type
            
            # Perform the reset
            results = self.order_controller.reset_order_manager(reset_type, self.user)
            
            if results["success"]:
                # Show success message with details
                total_archived = (results["active_archived"] + results["completed_archived"] + 
                                results["cancelled_archived"])
                
                message = f"Order manager reset completed successfully!\n\n"
                message += f"Orders archived:\n"
                message += f"• Active: {results['active_archived']}\n"
                message += f"• Completed: {results['completed_archived']}\n"
                message += f"• Cancelled: {results['cancelled_archived']}\n"
                message += f"• Old orders cleared: {results['old_cleared']}\n"
                message += f"• Total archived: {total_archived}"
                message += f"\n\n📋 Note: Order data is preserved."
                
                if results["errors"]:
                    message += f"\n\nWarnings: {len(results['errors'])} errors occurred during reset."
                
                QMessageBox.information(self, "Reset Complete", message)
            else:
                # Show error message
                error_message = "Failed to reset order manager:\n\n"
                for error in results["errors"]:
                    error_message += f"• {error}\n"
                
                QMessageBox.critical(self, "Reset Failed", error_message)
            
            # Refresh the order display
            self.refresh_orders()
            self.order_updated.emit()
    
    def __del__(self):
        """Clean up timer."""
        try:
            if hasattr(self, 'refresh_timer') and self.refresh_timer:
                self.refresh_timer.stop()
        except:
            pass  # Ignore errors during cleanup 


class DayReportDialog(QDialog):
    """Dialog for generating and viewing day reports."""
    
    def __init__(self, user: User, parent=None):
        super().__init__(parent)
        self.user = user
        self.order_controller = OrderController()
        self.report_data = None
        self.setWindowTitle("📊 Day Report Generator")
        self.setFixedSize(1400, 900)
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8f9fa, stop:1 #e9ecef);
            }
        """)
        self.init_ui()
    
    def init_ui(self):
        """Initialize the user interface."""
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(25, 25, 25, 25)
        
        # Header with gradient background
        header_frame = QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #667eea, stop:1 #764ba2);
                border-radius: 15px;
                padding: 20px;
            }
        """)
        header_frame.setFixedHeight(100)
        
        header_layout = QVBoxLayout(header_frame)
        header_layout.setContentsMargins(20, 15, 20, 15)
        
        # Title with icon and better styling
        title_layout = QHBoxLayout()
        title_layout.setSpacing(15)
        
        # Icon
        icon_label = QLabel("📊")
        icon_label.setStyleSheet("font-size: 40px;")
        title_layout.addWidget(icon_label)
        
        # Title
        title_label = QLabel("Day Report Generator")
        title_label.setFont(QFont("Arial", 24, QFont.Bold))
        title_label.setStyleSheet("color: white; margin: 0;")
        title_layout.addWidget(title_label)
        
        title_layout.addStretch()
        header_layout.addLayout(title_layout)
        
        # Subtitle
        subtitle_label = QLabel("Generate comprehensive analytics for any date")
        subtitle_label.setFont(QFont("Arial", 14))
        subtitle_label.setStyleSheet("color: rgba(255, 255, 255, 0.9); margin-top: 8px;")
        header_layout.addWidget(subtitle_label)
        
        layout.addWidget(header_frame)
        
        # Control panel with modern design
        control_frame = QFrame()
        control_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 12px;
                border: 1px solid #e0e0e0;
            }
        """)
        control_layout = QVBoxLayout(control_frame)
        control_layout.setContentsMargins(20, 20, 20, 20)
        control_layout.setSpacing(15)
        
        # Control panel title
        control_title = QLabel("📅 Report Configuration")
        control_title.setFont(QFont("Arial", 16, QFont.Bold))
        control_title.setStyleSheet("color: #2c3e50; margin-bottom: 15px;")
        control_layout.addWidget(control_title)
        
        # Date selection with better layout
        date_layout = QHBoxLayout()
        date_layout.setSpacing(15)
        
        # Date label with icon
        date_label = QLabel("🗓️ Select Date:")
        date_label.setFont(QFont("Arial", 14, QFont.Bold))
        date_label.setStyleSheet("color: #34495e; min-width: 140px;")
        date_layout.addWidget(date_label)
        
        # Date picker with enhanced styling
        self.date_edit = QDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setFixedHeight(50)
        self.date_edit.setStyleSheet("""
            QDateEdit {
                padding: 8px 12px;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                color: #2c3e50;
            }
            QDateEdit:focus {
                border-color: #667eea;
                background-color: #f8f9fa;
            }
            QDateEdit::drop-down {
                border: none;
                width: 20px;
            }
            QDateEdit::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #667eea;
                margin-right: 5px;
            }
        """)
        date_layout.addWidget(self.date_edit)
        
        date_layout.addStretch()
        control_layout.addLayout(date_layout)
        
        # Buttons with modern design
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        
        # Generate button with enhanced styling
        self.generate_btn = QPushButton("🚀 Generate Report")
        self.generate_btn.clicked.connect(self.generate_report)
        self.generate_btn.setFixedHeight(55)
        self.generate_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #667eea, stop:1 #764ba2);
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 24px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #5a6fd8, stop:1 #6a4190);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4a5fc8, stop:1 #5a3180);
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        button_layout.addWidget(self.generate_btn)
        
        # Export button with enhanced styling
        self.export_btn = QPushButton("📊 Export to Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setEnabled(False)
        self.export_btn.setFixedHeight(55)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #27ae60, stop:1 #2ecc71);
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 24px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #229954, stop:1 #27ae60);
            }
            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #1e8449, stop:1 #229954);
            }
            QPushButton:disabled {
                background: #bdc3c7;
                color: #7f8c8d;
            }
        """)
        button_layout.addWidget(self.export_btn)
        
        button_layout.addStretch()
        control_layout.addLayout(button_layout)
        
        layout.addWidget(control_frame)
        
        # Progress bar with modern styling
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setFixedHeight(8)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: none;
                border-radius: 4px;
                background-color: #ecf0f1;
                text-align: center;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #667eea, stop:1 #764ba2);
                border-radius: 4px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        # Report display area with enhanced styling
        self.tab_widget = QTabWidget()
        self.tab_widget.setVisible(False)
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #f8f9fa;
                color: #2c3e50;
                padding: 12px 20px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background-color: white;
                color: #667eea;
                border-bottom: 3px solid #667eea;
            }
            QTabBar::tab:hover {
                background-color: #e9ecef;
            }
        """)
        layout.addWidget(self.tab_widget)
        
        # Create tabs
        self.create_summary_tab()
        self.create_financial_tab()
        self.create_products_tab()
        self.create_users_tab()
        self.create_customers_tab()
        self.create_hourly_tab()
        self.create_orders_tab()
    
    def create_summary_tab(self):
        """Create the summary tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Tab header
        header_label = QLabel("📊 Summary Statistics")
        header_label.setFont(QFont("Arial", 16, QFont.Bold))
        header_label.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(header_label)
        
        # Summary cards
        self.summary_cards = {}
        summary_layout = QGridLayout()
        summary_layout.setSpacing(15)
        
        # Create summary cards
        cards = [
            ('total_orders', 'Total Orders', '#3498db'),
            ('completed_orders', 'Completed Orders', '#27ae60'),
            ('cancelled_orders', 'Cancelled Orders', '#e74c3c'),
            ('active_orders', 'Active Orders', '#f39c12'),
            ('completion_rate', 'Completion Rate (%)', '#9b59b6'),
            ('cancellation_rate', 'Cancellation Rate (%)', '#e67e22')
        ]
        
        for i, (key, title, color) in enumerate(cards):
            card = self.create_summary_card(title, "0", color)
            self.summary_cards[key] = card
            summary_layout.addWidget(card, i // 3, i % 3)
        
        layout.addLayout(summary_layout)
        layout.addStretch()
        
        self.tab_widget.addTab(tab, "📊 Summary")
    
    def create_financial_tab(self):
        """Create the financial tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Tab header
        header_label = QLabel("💰 Financial Analysis")
        header_label.setFont(QFont("Arial", 16, QFont.Bold))
        header_label.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(header_label)
        
        # Financial cards
        self.financial_cards = {}
        financial_layout = QGridLayout()
        financial_layout.setSpacing(15)
        
        # Create financial cards
        cards = [
            ('total_revenue', 'Total Revenue', '#27ae60'),
            ('total_subtotal', 'Total Subtotal', '#3498db'),
            ('total_tax', 'Total Tax', '#e67e22'),
            ('total_discount', 'Total Discount', '#e74c3c'),
            ('avg_order_value', 'Avg Order Value', '#9b59b6'),
            ('avg_subtotal', 'Avg Subtotal', '#f39c12')
        ]
        
        for i, (key, title, color) in enumerate(cards):
            card = self.create_summary_card(title, "$0.00", color)
            self.financial_cards[key] = card
            financial_layout.addWidget(card, i // 3, i % 3)
        
        layout.addLayout(financial_layout)
        layout.addStretch()
        
        self.tab_widget.addTab(tab, "💰 Financial")
    
    def create_products_tab(self):
        """Create the products tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Tab header
        header_label = QLabel("📦 Product Sales Analysis")
        header_label.setFont(QFont("Arial", 16, QFont.Bold))
        header_label.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(header_label)
        
        # Create tab widget for different views
        self.products_tab_widget = QTabWidget()
        self.products_tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #f8f9fa;
                color: #495057;
                padding: 12px 20px;
                margin-right: 2px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: bold;
                font-size: 13px;
            }
            QTabBar::tab:selected {
                background-color: #667eea;
                color: white;
                border-bottom: 3px solid #667eea;
            }
            QTabBar::tab:hover {
                background-color: #e9ecef;
            }
        """)
        
        # Summary view tab
        self.create_products_summary_tab()
        
        # Detailed sales view tab
        self.create_products_detailed_tab()
        
        layout.addWidget(self.products_tab_widget)
        
        self.tab_widget.addTab(tab, "📦 Products")
    
    def create_products_summary_tab(self):
        """Create the products summary tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Tab header
        header_label = QLabel("📊 Product Performance Summary")
        header_label.setFont(QFont("Arial", 14, QFont.Bold))
        header_label.setStyleSheet("color: #2c3e50; margin-bottom: 15px;")
        layout.addWidget(header_label)
        
        # Products summary table
        self.products_summary_table = QTableWidget()
        self.products_summary_table.setColumnCount(6)
        self.products_summary_table.setHorizontalHeaderLabels([
            'Product Name', 'Category', 'Quantity Sold', 'Total Revenue', 'Avg Price', 'Orders Count'
        ])
        self.products_summary_table.horizontalHeader().setStretchLastSection(True)
        self.products_summary_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                gridline-color: #f0f0f0;
                selection-background-color: #667eea;
                selection-color: white;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 12px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background-color: #667eea;
                color: white;
            }
            QHeaderView::section {
                background-color: #667eea;
                color: white;
                padding: 15px 10px;
                border: none;
                font-weight: bold;
                font-size: 14px;
            }
            QHeaderView::section:first {
                border-top-left-radius: 8px;
            }
            QHeaderView::section:last {
                border-top-right-radius: 8px;
            }
        """)
        layout.addWidget(self.products_summary_table)
        
        self.products_tab_widget.addTab(tab, "📊 Summary")
    
    def create_products_detailed_tab(self):
        """Create the products detailed sales tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Tab header
        header_label = QLabel("🛒 Individual Product Sales")
        header_label.setFont(QFont("Arial", 14, QFont.Bold))
        header_label.setStyleSheet("color: #2c3e50; margin-bottom: 15px;")
        layout.addWidget(header_label)
        
        # Products detailed table
        self.products_detailed_table = QTableWidget()
        self.products_detailed_table.setColumnCount(8)
        self.products_detailed_table.setHorizontalHeaderLabels([
            'Order #', 'Product Name', 'Category', 'Quantity', 'Unit Price', 'Total Price', 'Customer', 'Time'
        ])
        self.products_detailed_table.horizontalHeader().setStretchLastSection(True)
        self.products_detailed_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                gridline-color: #f0f0f0;
                selection-background-color: #667eea;
                selection-color: white;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 12px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background-color: #667eea;
                color: white;
            }
            QHeaderView::section {
                background-color: #667eea;
                color: white;
                padding: 15px 10px;
                border: none;
                font-weight: bold;
                font-size: 14px;
            }
            QHeaderView::section:first {
                border-top-left-radius: 8px;
            }
            QHeaderView::section:last {
                border-top-right-radius: 8px;
            }
        """)
        layout.addWidget(self.products_detailed_table)
        
        self.products_tab_widget.addTab(tab, "🛒 Detailed Sales")
    
    def create_users_tab(self):
        """Create the users tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Tab header
        header_label = QLabel("👥 User Performance Analysis")
        header_label.setFont(QFont("Arial", 16, QFont.Bold))
        header_label.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(header_label)
        
        # Users table with enhanced styling
        self.users_table = QTableWidget()
        self.users_table.setColumnCount(4)
        self.users_table.setHorizontalHeaderLabels([
            'User Name', 'Orders', 'Revenue', 'Avg Order Value'
        ])
        self.users_table.horizontalHeader().setStretchLastSection(True)
        self.users_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                gridline-color: #f0f0f0;
                selection-background-color: #667eea;
                selection-color: white;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 12px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background-color: #667eea;
                color: white;
            }
            QHeaderView::section {
                background-color: #667eea;
                color: white;
                padding: 15px 10px;
                border: none;
                font-weight: bold;
                font-size: 14px;
            }
            QHeaderView::section:first {
                border-top-left-radius: 8px;
            }
            QHeaderView::section:last {
                border-top-right-radius: 8px;
            }
        """)
        layout.addWidget(self.users_table)
        
        self.tab_widget.addTab(tab, "👥 Users")
    
    def create_customers_tab(self):
        """Create the customers tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Tab header
        header_label = QLabel("👤 Customer Analysis")
        header_label.setFont(QFont("Arial", 16, QFont.Bold))
        header_label.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(header_label)
        
        # Customers table with enhanced styling
        self.customers_table = QTableWidget()
        self.customers_table.setColumnCount(4)
        self.customers_table.setHorizontalHeaderLabels([
            'Customer Name', 'Orders', 'Total Spent', 'Avg Order Value'
        ])
        self.customers_table.horizontalHeader().setStretchLastSection(True)
        self.customers_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                gridline-color: #f0f0f0;
                selection-background-color: #667eea;
                selection-color: white;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 12px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background-color: #667eea;
                color: white;
            }
            QHeaderView::section {
                background-color: #667eea;
                color: white;
                padding: 15px 10px;
                border: none;
                font-weight: bold;
                font-size: 14px;
            }
            QHeaderView::section:first {
                border-top-left-radius: 8px;
            }
            QHeaderView::section:last {
                border-top-right-radius: 8px;
            }
        """)
        layout.addWidget(self.customers_table)
        
        self.tab_widget.addTab(tab, "👤 Customers")
    
    def create_hourly_tab(self):
        """Create the hourly tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Tab header
        header_label = QLabel("🕐 Hourly Distribution Analysis")
        header_label.setFont(QFont("Arial", 16, QFont.Bold))
        header_label.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(header_label)
        
        # Hourly table with enhanced styling
        self.hourly_table = QTableWidget()
        self.hourly_table.setColumnCount(4)
        self.hourly_table.setHorizontalHeaderLabels([
            'Hour', 'Orders', 'Revenue', 'Avg Order Value'
        ])
        self.hourly_table.horizontalHeader().setStretchLastSection(True)
        self.hourly_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                gridline-color: #f0f0f0;
                selection-background-color: #667eea;
                selection-color: white;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 12px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background-color: #667eea;
                color: white;
            }
            QHeaderView::section {
                background-color: #667eea;
                color: white;
                padding: 15px 10px;
                border: none;
                font-weight: bold;
                font-size: 14px;
            }
            QHeaderView::section:first {
                border-top-left-radius: 8px;
            }
            QHeaderView::section:last {
                border-top-right-radius: 8px;
            }
        """)
        layout.addWidget(self.hourly_table)
        
        self.tab_widget.addTab(tab, "🕐 Hourly")
    
    def create_orders_tab(self):
        """Create the orders tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        
        # Tab header
        header_label = QLabel("📋 Detailed Orders List")
        header_label.setFont(QFont("Arial", 16, QFont.Bold))
        header_label.setStyleSheet("color: #2c3e50; margin-bottom: 20px;")
        layout.addWidget(header_label)
        
        # Orders table with enhanced styling
        self.orders_table = QTableWidget()
        self.orders_table.setColumnCount(7)
        self.orders_table.setHorizontalHeaderLabels([
            'Order #', 'Customer', 'Status', 'Subtotal', 'Tax', 'Total', 'Created'
        ])
        self.orders_table.horizontalHeader().setStretchLastSection(True)
        self.orders_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                gridline-color: #f0f0f0;
                selection-background-color: #667eea;
                selection-color: white;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 12px;
                border-bottom: 1px solid #f0f0f0;
            }
            QTableWidget::item:selected {
                background-color: #667eea;
                color: white;
            }
            QHeaderView::section {
                background-color: #667eea;
                color: white;
                padding: 15px 10px;
                border: none;
                font-weight: bold;
                font-size: 14px;
            }
            QHeaderView::section:first {
                border-top-left-radius: 8px;
            }
            QHeaderView::section:last {
                border-top-right-radius: 8px;
            }
        """)
        layout.addWidget(self.orders_table)
        
        self.tab_widget.addTab(tab, "📋 Orders")
    
    def create_summary_card(self, title: str, value: str, color: str) -> QFrame:
        """Create a summary card widget with modern design."""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 white, stop:1 #f8f9fa);
                border: 2px solid {color};
                border-radius: 15px;
                padding: 20px;
                margin: 8px;
            }}
            QFrame:hover {{
                border-width: 3px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #ffffff, stop:1 #f0f0f0);
            }}
        """)
        card.setFixedHeight(140)
        
        layout = QVBoxLayout(card)
        layout.setSpacing(8)
        
        # Icon based on title
        icon_map = {
            'Total Orders': '📋',
            'Completed Orders': '✅',
            'Cancelled Orders': '❌',
            'Active Orders': '⏳',
            'Completion Rate': '📈',
            'Cancellation Rate': '📉',
            'Total Revenue': '💰',
            'Total Subtotal': '💵',
            'Total Tax': '🏛️',
            'Total Discount': '🎫',
            'Avg Order Value': '📊',
            'Avg Subtotal': '📈'
        }
        
        icon = icon_map.get(title, '📊')
        
        # Icon and title layout
        header_layout = QHBoxLayout()
        header_layout.setSpacing(10)
        
        # Icon
        icon_label = QLabel(icon)
        icon_label.setStyleSheet("font-size: 20px;")
        icon_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        header_layout.addWidget(icon_label)
        
        # Title
        title_label = QLabel(title)
        title_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 15px;")
        title_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        header_layout.addWidget(title_label)
        
        header_layout.addStretch()
        layout.addLayout(header_layout)
        
        # Value with enhanced styling
        value_label = QLabel(value)
        value_label.setStyleSheet(f"""
            font-weight: bold; 
            color: {color}; 
            font-size: 28px; 
            margin-top: 12px;
            padding: 8px;
            background-color: rgba(255, 255, 255, 0.7);
            border-radius: 10px;
        """)
        value_label.setAlignment(Qt.AlignCenter)
        value_label.setObjectName("value_label")  # Set object name for easy access
        layout.addWidget(value_label)
        
        return card
    
    def generate_report(self):
        """Generate the day report."""
        try:
            # Get selected date
            qdate = self.date_edit.date()
            target_date = date(qdate.year(), qdate.month(), qdate.day())
            
            # Show progress
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, 0)  # Indeterminate progress
            self.generate_btn.setEnabled(False)
            QApplication.processEvents()
            
            # Generate report
            self.report_data = self.order_controller.generate_day_report(target_date)
            
            if self.report_data:
                # Update UI with report data
                self.update_summary_tab()
                self.update_financial_tab()
                self.update_products_tab()
                self.update_users_tab()
                self.update_customers_tab()
                self.update_hourly_tab()
                self.update_orders_tab()
                
                # Show report
                self.tab_widget.setVisible(True)
                self.export_btn.setEnabled(True)
                
                QMessageBox.information(
                    self,
                    "Report Generated",
                    f"Day report for {target_date.strftime('%Y-%m-%d')} generated successfully!\n\n"
                    f"Total Orders: {self.report_data['summary']['total_orders']}\n"
                    f"Total Revenue: ${self.report_data['financial']['total_revenue']:.2f}"
                )
            else:
                QMessageBox.warning(
                    self,
                    "No Data",
                    f"No order data found for {target_date.strftime('%Y-%m-%d')}."
                )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error generating report: {str(e)}"
            )
        finally:
            self.progress_bar.setVisible(False)
            self.generate_btn.setEnabled(True)
    
    def update_summary_tab(self):
        """Update the summary tab with report data."""
        if not self.report_data:
            return
        
        summary = self.report_data['summary']
        
        # Update summary cards
        self.summary_cards['total_orders'].findChild(QLabel, "value_label").setText(str(summary['total_orders']))
        self.summary_cards['completed_orders'].findChild(QLabel, "value_label").setText(str(summary['completed_orders']))
        self.summary_cards['cancelled_orders'].findChild(QLabel, "value_label").setText(str(summary['cancelled_orders']))
        self.summary_cards['active_orders'].findChild(QLabel, "value_label").setText(str(summary['active_orders']))
        self.summary_cards['completion_rate'].findChild(QLabel, "value_label").setText(f"{summary['completion_rate']}%")
        self.summary_cards['cancellation_rate'].findChild(QLabel, "value_label").setText(f"{summary['cancellation_rate']}%")
    
    def update_financial_tab(self):
        """Update the financial tab with report data."""
        if not self.report_data:
            return
        
        financial = self.report_data['financial']
        
        # Update financial cards
        self.financial_cards['total_revenue'].findChild(QLabel, "value_label").setText(f"${financial['total_revenue']:.2f}")
        self.financial_cards['total_subtotal'].findChild(QLabel, "value_label").setText(f"${financial['total_subtotal']:.2f}")
        self.financial_cards['total_tax'].findChild(QLabel, "value_label").setText(f"${financial['total_tax']:.2f}")
        self.financial_cards['total_discount'].findChild(QLabel, "value_label").setText(f"${financial['total_discount']:.2f}")
        self.financial_cards['avg_order_value'].findChild(QLabel, "value_label").setText(f"${financial['avg_order_value']:.2f}")
        self.financial_cards['avg_subtotal'].findChild(QLabel, "value_label").setText(f"${financial['avg_subtotal']:.2f}")
    
    def update_products_tab(self):
        """Update the products tab with report data."""
        if not self.report_data:
            return
        
        # Update summary table
        self.update_products_summary_tab()
        
        # Update detailed sales table
        self.update_products_detailed_tab()
    
    def update_products_summary_tab(self):
        """Update the products summary tab with aggregated data."""
        if not self.report_data:
            return
        
        products = self.report_data['products']['product_statistics']
        
        self.products_summary_table.setRowCount(len(products))
        for i, product in enumerate(products):
            self.products_summary_table.setItem(i, 0, QTableWidgetItem(product['product_name']))
            self.products_summary_table.setItem(i, 1, QTableWidgetItem(product['category']))
            self.products_summary_table.setItem(i, 2, QTableWidgetItem(str(product['quantity'])))
            self.products_summary_table.setItem(i, 3, QTableWidgetItem(f"${product['total_revenue']:.2f}"))
            self.products_summary_table.setItem(i, 4, QTableWidgetItem(f"${product['avg_price']:.2f}"))
            self.products_summary_table.setItem(i, 5, QTableWidgetItem(str(product['orders_count'])))
    
    def update_products_detailed_tab(self):
        """Update the products detailed tab with individual sales data."""
        if not self.report_data:
            return
        
        # Get all completed orders for the day
        completed_orders = self.report_data['orders']['completed_orders']
        
        # Collect all product sales from all orders
        detailed_sales = []
        for order in completed_orders:
            order_items = order.get_order_items()
            for item in order_items:
                product = item['product']
                
                # Get category name safely
                category_name = 'Unknown'
                try:
                    if product.category:
                        category_name = product.category.name
                except Exception:
                    # If there's a session issue, try to get category from database
                    try:
                        from models.product import Product
                        from database.db_config import get_fresh_session
                        session = get_fresh_session()
                        try:
                            fresh_product = session.query(Product).filter_by(id=product.id).first()
                            if fresh_product and fresh_product.category:
                                category_name = fresh_product.category.name
                        finally:
                            session.close()
                    except Exception:
                        category_name = 'Unknown'
                
                detailed_sales.append({
                    'order_number': order.order_number,
                    'product_name': product.name,
                    'category': category_name,
                    'quantity': item['quantity'],
                    'unit_price': item['price'],
                    'total_price': item['price'] * item['quantity'],
                    'customer_name': order.customer_name or 'Anonymous',
                    'time': order.created_at.strftime('%H:%M:%S')
                })
        
        # Sort by time (most recent first)
        detailed_sales.sort(key=lambda x: x['time'], reverse=True)
        
        self.products_detailed_table.setRowCount(len(detailed_sales))
        for i, sale in enumerate(detailed_sales):
            self.products_detailed_table.setItem(i, 0, QTableWidgetItem(sale['order_number']))
            self.products_detailed_table.setItem(i, 1, QTableWidgetItem(sale['product_name']))
            self.products_detailed_table.setItem(i, 2, QTableWidgetItem(sale['category']))
            self.products_detailed_table.setItem(i, 3, QTableWidgetItem(str(sale['quantity'])))
            self.products_detailed_table.setItem(i, 4, QTableWidgetItem(f"${sale['unit_price']:.2f}"))
            self.products_detailed_table.setItem(i, 5, QTableWidgetItem(f"${sale['total_price']:.2f}"))
            self.products_detailed_table.setItem(i, 6, QTableWidgetItem(sale['customer_name']))
            self.products_detailed_table.setItem(i, 7, QTableWidgetItem(sale['time']))
    
    def update_users_tab(self):
        """Update the users tab with report data."""
        if not self.report_data:
            return
        
        users = self.report_data['users']['user_statistics']
        
        self.users_table.setRowCount(len(users))
        for i, user in enumerate(users):
            self.users_table.setItem(i, 0, QTableWidgetItem(user['user_name']))
            self.users_table.setItem(i, 1, QTableWidgetItem(str(user['orders_count'])))
            self.users_table.setItem(i, 2, QTableWidgetItem(f"${user['total_revenue']:.2f}"))
            self.users_table.setItem(i, 3, QTableWidgetItem(f"${user['avg_order_value']:.2f}"))
    
    def update_customers_tab(self):
        """Update the customers tab with report data."""
        if not self.report_data:
            return
        
        customers = self.report_data['customers']['customer_statistics']
        
        self.customers_table.setRowCount(len(customers))
        for i, customer in enumerate(customers):
            self.customers_table.setItem(i, 0, QTableWidgetItem(customer['customer_name']))
            self.customers_table.setItem(i, 1, QTableWidgetItem(str(customer['orders_count'])))
            self.customers_table.setItem(i, 2, QTableWidgetItem(f"${customer['total_spent']:.2f}"))
            self.customers_table.setItem(i, 3, QTableWidgetItem(f"${customer['avg_order_value']:.2f}"))
    
    def update_hourly_tab(self):
        """Update the hourly tab with report data."""
        if not self.report_data:
            return
        
        hourly = self.report_data['hourly']['hourly_statistics']
        
        self.hourly_table.setRowCount(len(hourly))
        for i, hour_data in enumerate(hourly):
            self.hourly_table.setItem(i, 0, QTableWidgetItem(f"{hour_data['hour']:02d}:00"))
            self.hourly_table.setItem(i, 1, QTableWidgetItem(str(hour_data['orders_count'])))
            self.hourly_table.setItem(i, 2, QTableWidgetItem(f"${hour_data['total_revenue']:.2f}"))
            self.hourly_table.setItem(i, 3, QTableWidgetItem(f"${hour_data['avg_order_value']:.2f}"))
    
    def update_orders_tab(self):
        """Update the orders tab with report data."""
        if not self.report_data:
            return
        
        orders = self.report_data['orders']['all_orders']
        
        self.orders_table.setRowCount(len(orders))
        for i, order in enumerate(orders):
            self.orders_table.setItem(i, 0, QTableWidgetItem(order.order_number))
            self.orders_table.setItem(i, 1, QTableWidgetItem(order.customer_name or 'Anonymous'))
            self.orders_table.setItem(i, 2, QTableWidgetItem(order.get_status_display()))
            self.orders_table.setItem(i, 3, QTableWidgetItem(f"${order.subtotal:.2f}"))
            self.orders_table.setItem(i, 4, QTableWidgetItem(f"${order.tax_amount:.2f}"))
            self.orders_table.setItem(i, 5, QTableWidgetItem(f"${order.total_amount:.2f}"))
            self.orders_table.setItem(i, 6, QTableWidgetItem(order.created_at.strftime('%H:%M:%S')))
    
    def export_to_excel(self):
        """Export the report to Excel."""
        if not self.report_data:
            QMessageBox.warning(self, "No Data", "No report data to export.")
            return
        
        try:
            from PyQt5.QtWidgets import QFileDialog
            
            # Check if Excel functionality is available
            if not EXCEL_AVAILABLE:
                QMessageBox.warning(
                    self,
                    "Excel Not Available",
                    "Excel functionality is not available. Please install openpyxl: pip install openpyxl"
                )
                return
            
            # Get save location
            qdate = self.date_edit.date()
            target_date = date(qdate.year(), qdate.month(), qdate.day())
            default_name = f"day_report_{target_date.strftime('%Y%m%d')}.xlsx"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Day Report",
                default_name,
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if file_path:
                # Generate Excel report
                success = self.generate_excel_report(file_path)
                
                if success:
                    QMessageBox.information(
                        self,
                        "Export Successful",
                        f"Day report exported to:\n{file_path}"
                    )
                else:
                    QMessageBox.critical(
                        self,
                        "Export Failed",
                        "Failed to export day report to Excel."
                    )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Error exporting to Excel: {str(e)}"
            )
    
    def generate_excel_report(self, file_path: str) -> bool:
        """Generate Excel report for the day."""
        try:
            from utils.excel_report_generator import ExcelReportGenerator
            import openpyxl
            
            # Check if Excel functionality is available
            if not EXCEL_AVAILABLE:
                logger.error("Excel functionality not available - openpyxl not installed")
                return False
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Create summary sheet
            self.create_excel_summary_sheet(wb, "Summary")
            
            # Create detailed sheets
            self.create_excel_products_sheet(wb, "Products")
            self.create_excel_users_sheet(wb, "Users")
            self.create_excel_customers_sheet(wb, "Customers")
            self.create_excel_hourly_sheet(wb, "Hourly")
            self.create_excel_orders_sheet(wb, "Orders")
            
            # Save workbook
            wb.save(file_path)
            wb.close()
            
            return True
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return False
    
    def create_excel_summary_sheet(self, wb, sheet_name: str):
        """Create summary sheet in Excel workbook."""
        ws = wb.create_sheet(sheet_name)
        
        # Header
        ws['A1'] = "TALINDA POS SYSTEM - DAY REPORT"
        ws['A1'].font = Font(bold=True, size=16, color="1976D2")
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date
        qdate = self.date_edit.date()
        target_date = date(qdate.year(), qdate.month(), qdate.day())
        ws['A2'] = f"Date: {target_date.strftime('%Y-%m-%d')}"
        ws['A2'].font = Font(bold=True, size=12)
        
        # Summary data
        row = 4
        summary = self.report_data['summary']
        financial = self.report_data['financial']
        
        summary_data = [
            ("Total Orders", summary['total_orders']),
            ("Completed Orders", summary['completed_orders']),
            ("Cancelled Orders", summary['cancelled_orders']),
            ("Active Orders", summary['active_orders']),
            ("Completion Rate", f"{summary['completion_rate']}%"),
            ("Cancellation Rate", f"{summary['cancellation_rate']}%"),
            ("", ""),
            ("Total Revenue", f"${financial['total_revenue']:.2f}"),
            ("Total Subtotal", f"${financial['total_subtotal']:.2f}"),
            ("Total Tax", f"${financial['total_tax']:.2f}"),
            ("Total Discount", f"${financial['total_discount']:.2f}"),
            ("Average Order Value", f"${financial['avg_order_value']:.2f}"),
            ("Average Subtotal", f"${financial['avg_subtotal']:.2f}")
        ]
        
        for title, value in summary_data:
            ws[f'A{row}'] = title
            ws[f'B{row}'] = value
            if title:  # Skip empty rows
                ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def create_excel_products_sheet(self, wb, sheet_name: str):
        """Create products sheet in Excel workbook."""
        ws = wb.create_sheet(sheet_name)
        
        # Summary section
        ws.cell(row=1, column=1, value="PRODUCT PERFORMANCE SUMMARY").font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Summary headers
        summary_headers = ['Product Name', 'Category', 'Quantity Sold', 'Total Revenue', 'Avg Price', 'Orders Count']
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Summary data
        products = self.report_data['products']['product_statistics']
        for row, product in enumerate(products, 4):
            ws.cell(row=row, column=1, value=product['product_name'])
            ws.cell(row=row, column=2, value=product['category'])
            ws.cell(row=row, column=3, value=product['quantity'])
            ws.cell(row=row, column=4, value=product['total_revenue'])
            ws.cell(row=row, column=5, value=product['avg_price'])
            ws.cell(row=row, column=6, value=product['orders_count'])
        
        # Add space before detailed section
        start_row = len(products) + 6
        
        # Detailed section
        ws.cell(row=start_row, column=1, value="DETAILED PRODUCT SALES").font = Font(bold=True, size=14)
        ws.merge_cells(f'A{start_row}:H{start_row}')
        
        # Detailed headers
        detailed_headers = ['Order #', 'Product Name', 'Category', 'Quantity', 'Unit Price', 'Total Price', 'Customer', 'Time']
        for col, header in enumerate(detailed_headers, 1):
            cell = ws.cell(row=start_row + 2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Detailed data
        completed_orders = self.report_data['orders']['completed_orders']
        detailed_sales = []
        for order in completed_orders:
            order_items = order.get_order_items()
            for item in order_items:
                product = item['product']
                
                # Get category name safely
                category_name = 'Unknown'
                try:
                    if product.category:
                        category_name = product.category.name
                except Exception:
                    # If there's a session issue, try to get category from database
                    try:
                        from models.product import Product
                        from database.db_config import get_fresh_session
                        session = get_fresh_session()
                        try:
                            fresh_product = session.query(Product).filter_by(id=product.id).first()
                            if fresh_product and fresh_product.category:
                                category_name = fresh_product.category.name
                        finally:
                            session.close()
                    except Exception:
                        category_name = 'Unknown'
                
                detailed_sales.append({
                    'order_number': order.order_number,
                    'product_name': product.name,
                    'category': category_name,
                    'quantity': item['quantity'],
                    'unit_price': item['price'],
                    'total_price': item['price'] * item['quantity'],
                    'customer_name': order.customer_name or 'Anonymous',
                    'time': order.created_at.strftime('%H:%M:%S')
                })
        
        # Sort by time (most recent first)
        detailed_sales.sort(key=lambda x: x['time'], reverse=True)
        
        for row, sale in enumerate(detailed_sales, start_row + 3):
            ws.cell(row=row, column=1, value=sale['order_number'])
            ws.cell(row=row, column=2, value=sale['product_name'])
            ws.cell(row=row, column=3, value=sale['category'])
            ws.cell(row=row, column=4, value=sale['quantity'])
            ws.cell(row=row, column=5, value=sale['unit_price'])
            ws.cell(row=row, column=6, value=sale['total_price'])
            ws.cell(row=row, column=7, value=sale['customer_name'])
            ws.cell(row=row, column=8, value=sale['time'])
    
    def create_excel_users_sheet(self, wb, sheet_name: str):
        """Create users sheet in Excel workbook."""
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = ['User Name', 'Orders', 'Revenue', 'Avg Order Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        users = self.report_data['users']['user_statistics']
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1, value=user['user_name'])
            ws.cell(row=row, column=2, value=user['orders_count'])
            ws.cell(row=row, column=3, value=user['total_revenue'])
            ws.cell(row=row, column=4, value=user['avg_order_value'])
    
    def create_excel_customers_sheet(self, wb, sheet_name: str):
        """Create customers sheet in Excel workbook."""
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = ['Customer Name', 'Orders', 'Total Spent', 'Avg Order Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        customers = self.report_data['customers']['customer_statistics']
        for row, customer in enumerate(customers, 2):
            ws.cell(row=row, column=1, value=customer['customer_name'])
            ws.cell(row=row, column=2, value=customer['orders_count'])
            ws.cell(row=row, column=3, value=customer['total_spent'])
            ws.cell(row=row, column=4, value=customer['avg_order_value'])
    
    def create_excel_hourly_sheet(self, wb, sheet_name: str):
        """Create hourly sheet in Excel workbook."""
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = ['Hour', 'Orders', 'Revenue', 'Avg Order Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        hourly = self.report_data['hourly']['hourly_statistics']
        for row, hour_data in enumerate(hourly, 2):
            ws.cell(row=row, column=1, value=f"{hour_data['hour']:02d}:00")
            ws.cell(row=row, column=2, value=hour_data['orders_count'])
            ws.cell(row=row, column=3, value=hour_data['total_revenue'])
            ws.cell(row=row, column=4, value=hour_data['avg_order_value'])
    
    def create_excel_orders_sheet(self, wb, sheet_name: str):
        """Create orders sheet in Excel workbook."""
        ws = wb.create_sheet(sheet_name)
        
        # Headers
        headers = ['Order #', 'Customer', 'Status', 'Subtotal', 'Tax', 'Total', 'Created']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        orders = self.report_data['orders']['all_orders']
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order.order_number)
            ws.cell(row=row, column=2, value=order.customer_name or 'Anonymous')
            ws.cell(row=row, column=3, value=order.get_status_display())
            ws.cell(row=row, column=4, value=order.subtotal)
            ws.cell(row=row, column=5, value=order.tax_amount)
            ws.cell(row=row, column=6, value=order.total_amount)
            ws.cell(row=row, column=7, value=order.created_at.strftime('%H:%M:%S')) 