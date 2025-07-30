#!/usr/bin/env python3
"""
Excel Report Generator for Talinda POS System
============================================

Generates comprehensive Excel reports for shift summaries including:
- Opening amount
- All product sales during the shift
- Closing amount
- Sales totals and statistics

Author: Talinda POS Team
Version: 1.0.0
"""

import os
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Union
from sqlalchemy.orm import Session
from database.db_config import get_fresh_session
from models.user import Shift, User
from models.sale import Sale, sale_products
from models.product import Product
from sqlalchemy import func, and_

logger = logging.getLogger(__name__)

# Define a dummy Worksheet class for type hints when openpyxl is not available
class Worksheet:
    pass

# Global variables to store Excel imports
_openpyxl_available = False
_Font = None
_PatternFill = None
_Alignment = None
_Border = None
_Side = None
_get_column_letter = None

def check_excel_availability():
    """Check if openpyxl is available and return availability status."""
    global _openpyxl_available, _Font, _PatternFill, _Alignment, _Border, _Side, _get_column_letter
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.worksheet import Worksheet
        
        # Store the imports globally
        _Font = Font
        _PatternFill = PatternFill
        _Alignment = Alignment
        _Border = Border
        _Side = Side
        _get_column_letter = get_column_letter
        _openpyxl_available = True
        
        logger.info("openpyxl successfully imported - Excel functionality available")
        return True
    except ImportError as e:
        logger.error(f"openpyxl import failed: {e}. Excel reports will not be generated.")
        logger.error("To fix this, install openpyxl: pip install openpyxl")
        _openpyxl_available = False
        return False
    except Exception as e:
        logger.error(f"Unexpected error importing openpyxl: {e}. Excel reports will not be generated.")
        _openpyxl_available = False
        return False

# Check Excel availability at module load time
EXCEL_AVAILABLE = check_excel_availability()


class ExcelReportGenerator:
    """Generates Excel reports for shift summaries."""
    
    def __init__(self):
        """Initialize the Excel report generator."""
        self.session = get_fresh_session()
        self.reports_dir = Path("reports")
        self.reports_dir.mkdir(exist_ok=True)
        
        # Excel styling constants - only initialize if openpyxl is available
        if self.is_excel_available():
            try:
                self.header_font = _Font(bold=True, size=12, color="FFFFFF")
                self.header_fill = _PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
                self.subheader_font = _Font(bold=True, size=11, color="2C3E50")
                self.subheader_fill = _PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
                self.border = _Border(
                    left=_Side(style='thin'),
                    right=_Side(style='thin'),
                    top=_Side(style='thin'),
                    bottom=_Side(style='thin')
                )
                logger.debug("Excel styling constants initialized successfully")
            except Exception as e:
                logger.error(f"Error initializing Excel styling constants: {e}")
                # Set to None if styling initialization fails
                self.header_font = None
                self.header_fill = None
                self.subheader_font = None
                self.subheader_fill = None
                self.border = None
        else:
            # Set to None when openpyxl is not available
            self.header_font = None
            self.header_fill = None
            self.subheader_font = None
            self.subheader_fill = None
            self.border = None
            logger.warning("Excel styling constants not initialized - openpyxl not available")
    
    def generate_shift_report(self, shift: Shift) -> Optional[str]:
        """
        Generate a comprehensive Excel report for a shift.
        
        Args:
            shift: The shift object containing shift data
            
        Returns:
            Optional[str]: Path to the generated Excel file, or None if failed
        """
        if not self.is_excel_available():
            logger.error("Excel functionality not available. Install openpyxl: pip install openpyxl")
            return None
        
        # Additional check to ensure openpyxl is properly imported
        try:
            import openpyxl
            logger.debug("openpyxl import check passed")
        except ImportError as e:
            logger.error(f"openpyxl import failed during report generation: {e}")
            logger.error("Please reinstall openpyxl: pip install openpyxl")
            return None
        except Exception as e:
            logger.error(f"Unexpected error during openpyxl import check: {e}")
            return None
        
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Shift Report"
            
            # Generate report content
            self._create_header(ws, shift)
            self._create_shift_summary(ws, shift)
            self._create_sales_summary(ws, shift)
            self._create_detailed_sales(ws, shift)
            self._create_product_summary(ws, shift)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            # Save the file
            filename = self._generate_filename(shift)
            filepath = self.reports_dir / filename
            
            # Ensure the reports directory exists
            self.reports_dir.mkdir(exist_ok=True)
            
            # Save with error handling
            try:
                wb.save(str(filepath))
                wb.close()  # Explicitly close the workbook
                logger.info(f"Shift report generated: {filepath}")
                return str(filepath)
            except PermissionError as e:
                logger.error(f"Permission error saving report: {e}")
                # Try with a different filename
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                alt_filename = f"shift_report_{shift.user.username}_{timestamp}.xlsx"
                alt_filepath = self.reports_dir / alt_filename
                wb.save(str(alt_filepath))
                wb.close()
                logger.info(f"Shift report generated with alternate filename: {alt_filepath}")
                return str(alt_filepath)
            except Exception as e:
                logger.error(f"Error saving workbook: {e}")
                wb.close()
                return None
            
        except Exception as e:
            logger.error(f"Error generating shift report: {str(e)}")
            return None
    
    def _create_header(self, ws: Union[Worksheet, object], shift: Shift):
        """Create the report header."""
        if not EXCEL_AVAILABLE:
            return
            
        # Company title
        ws['A1'] = "TALINDA POS SYSTEM"
        ws['A1'].font = _Font(bold=True, size=16, color="1976D2")
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = _Alignment(horizontal='center')
        
        # Report title
        ws['A2'] = "SHIFT SUMMARY REPORT"
        ws['A2'].font = _Font(bold=True, size=14, color="2C3E50")
        ws.merge_cells('A2:H2')
        ws['A2'].alignment = _Alignment(horizontal='center')
        
        # Shift information
        ws['A4'] = "Shift Details:"
        if self.subheader_font:
            ws['A4'].font = self.subheader_font
        
        ws['A5'] = f"Cashier: {shift.user.full_name or shift.user.username}"
        ws['B5'] = f"User ID: {shift.user_id}"
        
        ws['A6'] = f"Shift ID: {shift.id}"
        ws['B6'] = f"Status: {shift.status.value.upper()}"
        
        ws['A7'] = f"Open Time: {shift.open_time.strftime('%Y-%m-%d %I:%M:%S %p')}"
        if shift.close_time:
            ws['B7'] = f"Close Time: {shift.close_time.strftime('%Y-%m-%d %I:%M:%S %p')}"
        
        # Report generation time
        ws['A9'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')}"
        ws['A9'].font = _Font(italic=True, size=10, color="7F8C8D")
    
    def _create_shift_summary(self, ws: Union[Worksheet, object], shift: Shift):
        """Create the shift summary section."""
        if not EXCEL_AVAILABLE:
            return
            
        # Section header
        row = 12
        ws[f'A{row}'] = "SHIFT SUMMARY"
        if self.header_font:
            ws[f'A{row}'].font = self.header_font
        if self.header_fill:
            ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].alignment = _Alignment(horizontal='center')
        
        # Summary data
        row += 2
        ws[f'A{row}'] = "Opening Amount:"
        ws[f'B{row}'] = f"${shift.opening_amount:.2f}"
        ws[f'B{row}'].font = _Font(bold=True, color="27AE60")
        
        # Calculate total sales
        total_sales = self._get_shift_total_sales(shift)
        row += 1
        ws[f'A{row}'] = "Total Sales:"
        ws[f'B{row}'] = f"${total_sales:.2f}"
        ws[f'B{row}'].font = _Font(bold=True, color="3498DB")
        
        # Calculate expected cash
        expected_cash = shift.opening_amount + total_sales
        row += 1
        ws[f'A{row}'] = "Expected Cash:"
        ws[f'B{row}'] = f"${expected_cash:.2f}"
        ws[f'B{row}'].font = _Font(bold=True, color="F39C12")
    
    def _create_sales_summary(self, ws: Union[Worksheet, object], shift: Shift):
        """Create the sales summary section."""
        if not EXCEL_AVAILABLE:
            return
            
        # Section header
        row = 22
        ws[f'A{row}'] = "SALES SUMMARY"
        if self.header_font:
            ws[f'A{row}'].font = self.header_font
        if self.header_fill:
            ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].alignment = _Alignment(horizontal='center')
        
        # Get sales data
        sales_data = self._get_shift_sales_data(shift)
        
        # Sales statistics
        row += 2
        ws[f'A{row}'] = "Total Transactions:"
        ws[f'B{row}'] = len(sales_data)
        ws[f'B{row}'].font = _Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Total Items Sold:"
        total_items = sum(sale['total_items'] for sale in sales_data)
        ws[f'B{row}'] = total_items
        ws[f'B{row}'].font = _Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Average Transaction:"
        if sales_data:
            avg_transaction = sum(sale['total_amount'] for sale in sales_data) / len(sales_data)
            ws[f'B{row}'] = f"${avg_transaction:.2f}"
            ws[f'B{row}'].font = _Font(bold=True)
    
    def _create_detailed_sales(self, ws: Union[Worksheet, object], shift: Shift):
        """Create the detailed sales section."""
        if not EXCEL_AVAILABLE:
            return
            
        # Section header
        row = 30
        ws[f'A{row}'] = "DETAILED SALES"
        if self.header_font:
            ws[f'A{row}'].font = self.header_font
        if self.header_fill:
            ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].alignment = _Alignment(horizontal='center')
        
        # Table headers
        row += 2
        headers = ['Sale ID', 'Time', 'Items', 'Total Amount', 'Cashier']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            if self.subheader_font:
                cell.font = self.subheader_font
            if self.subheader_fill:
                cell.fill = self.subheader_fill
            if self.border:
                cell.border = self.border
            cell.alignment = _Alignment(horizontal='center')
        
        # Get sales data
        sales_data = self._get_shift_sales_data(shift)
        
        # Sales rows
        for sale in sales_data:
            row += 1
            ws.cell(row=row, column=1, value=sale['id'])
            ws.cell(row=row, column=2, value=sale['timestamp'].strftime('%I:%M:%S %p'))
            ws.cell(row=row, column=3, value=sale['total_items'])
            ws.cell(row=row, column=4, value=f"${sale['total_amount']:.2f}")
            ws.cell(row=row, column=5, value=sale['cashier_name'])
            
            # Apply borders to all cells in the row
            if self.border:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = self.border
    
    def _create_product_summary(self, ws: Union[Worksheet, object], shift: Shift):
        """Create the product summary section."""
        if not EXCEL_AVAILABLE:
            return
            
        # Section header
        row = 35 + len(self._get_shift_sales_data(shift))  # Start after detailed sales
        ws[f'A{row}'] = "PRODUCT SUMMARY"
        if self.header_font:
            ws[f'A{row}'].font = self.header_font
        if self.header_fill:
            ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].alignment = _Alignment(horizontal='center')
        
        # Table headers
        row += 2
        headers = ['Product Name', 'Category', 'Quantity Sold', 'Unit Price', 'Total Revenue']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            if self.subheader_font:
                cell.font = self.subheader_font
            if self.subheader_fill:
                cell.fill = self.subheader_fill
            if self.border:
                cell.border = self.border
            cell.alignment = _Alignment(horizontal='center')
        
        # Get product sales data
        product_data = self._get_shift_product_data(shift)
        
        # Product rows
        for product in product_data:
            row += 1
            ws.cell(row=row, column=1, value=product['name'])
            ws.cell(row=row, column=2, value=product['category'])
            ws.cell(row=row, column=3, value=product['quantity'])
            ws.cell(row=row, column=4, value=f"${product['unit_price']:.2f}")
            ws.cell(row=row, column=5, value=f"${product['total_revenue']:.2f}")
            
            # Apply borders to all cells in the row
            if self.border:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = self.border
    
    def _get_shift_total_sales(self, shift: Shift) -> float:
        """Get total sales amount for the shift."""
        try:
            result = self.session.query(func.sum(Sale.total_amount)).filter(
                and_(
                    Sale.user_id == shift.user_id,
                    Sale.timestamp >= shift.open_time,
                    Sale.timestamp <= (shift.close_time or datetime.now())
                )
            ).scalar()
            return float(result or 0)
        except Exception as e:
            logger.error(f"Error getting shift total sales: {e}")
            return 0.0
    
    def _get_shift_sales_data(self, shift: Shift) -> List[Dict]:
        """Get detailed sales data for the shift."""
        try:
            sales = self.session.query(Sale).filter(
                and_(
                    Sale.user_id == shift.user_id,
                    Sale.timestamp >= shift.open_time,
                    Sale.timestamp <= (shift.close_time or datetime.now())
                )
            ).order_by(Sale.timestamp).all()
            
            sales_data = []
            for sale in sales:
                # Count total items in this sale
                total_items = self.session.query(func.sum(sale_products.c.quantity)).filter(
                    sale_products.c.sale_id == sale.id
                ).scalar() or 0
                
                sales_data.append({
                    'id': sale.id,
                    'timestamp': sale.timestamp,
                    'total_amount': sale.total_amount,
                    'total_items': total_items,
                    'cashier_name': shift.user.full_name or shift.user.username
                })
            
            return sales_data
        except Exception as e:
            logger.error(f"Error getting shift sales data: {e}")
            return []
    
    def _get_shift_product_data(self, shift: Shift) -> List[Dict]:
        """Get product sales data for the shift."""
        try:
            # Get all sales for this shift
            sales = self.session.query(Sale.id).filter(
                and_(
                    Sale.user_id == shift.user_id,
                    Sale.timestamp >= shift.open_time,
                    Sale.timestamp <= (shift.close_time or datetime.now())
                )
            ).all()
            
            sale_ids = [sale.id for sale in sales]
            
            if not sale_ids:
                return []
            
            # Get product sales data
            product_sales = self.session.query(
                Product.name,
                Product.category_id,
                func.sum(sale_products.c.quantity).label('total_quantity'),
                func.avg(sale_products.c.price_at_sale).label('avg_price'),
                func.sum(sale_products.c.quantity * sale_products.c.price_at_sale).label('total_revenue')
            ).join(sale_products, Product.id == sale_products.c.product_id).filter(
                sale_products.c.sale_id.in_(sale_ids)
            ).group_by(Product.id, Product.name, Product.category_id).all()
            
            # Get category names
            category_names = {}
            categories = self.session.query(Product.category_id, func.count(Product.id)).group_by(Product.category_id).all()
            for cat_id, _ in categories:
                if cat_id:
                    category = self.session.query(Product).filter_by(category_id=cat_id).first()
                    if category and category.category:
                        category_names[cat_id] = category.category.name
                    else:
                        category_names[cat_id] = "Unknown"
            
            product_data = []
            for product in product_sales:
                product_data.append({
                    'name': product.name,
                    'category': category_names.get(product.category_id, "Unknown"),
                    'quantity': int(product.total_quantity),
                    'unit_price': float(product.avg_price),
                    'total_revenue': float(product.total_revenue)
                })
            
            return product_data
        except Exception as e:
            logger.error(f"Error getting shift product data: {e}")
            return []
    
    def _auto_adjust_columns(self, ws: Union[Worksheet, object]):
        """Auto-adjust column widths based on content."""
        if not EXCEL_AVAILABLE:
            return
            
        for column in ws.columns:
            max_length = 0
            column_letter = _get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _generate_filename(self, shift: Shift) -> str:
        """Generate filename for the Excel report."""
        timestamp = shift.close_time or datetime.now()
        return f"shift_report_{shift.user.username}_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    def open_excel_file(self, filepath: str) -> bool:
        """
        Open the Excel file with the default system application.
        
        Args:
            filepath: Path to the Excel file
            
        Returns:
            bool: True if opened successfully, False otherwise
        """
        try:
            import subprocess
            import platform
            
            system = platform.system()
            
            if system == "Windows":
                os.startfile(filepath)
            elif system == "Darwin":  # macOS
                subprocess.run(["open", filepath], check=True)
            else:  # Linux
                subprocess.run(["xdg-open", filepath], check=True)
            
            logger.info(f"Excel file opened: {filepath}")
            return True
            
        except Exception as e:
            logger.error(f"Error opening Excel file: {e}")
            return False
    
    def save_report_as(self, shift: Shift, parent_widget=None) -> Optional[str]:
        """
        Generate and save Excel report to a user-specified location.
        
        Args:
            shift: The shift object containing shift data
            parent_widget: Parent widget for file dialog (optional)
            
        Returns:
            Optional[str]: Path to the saved Excel file, or None if failed
        """
        try:
            from PyQt5.QtWidgets import QFileDialog
            from pathlib import Path
            from datetime import datetime
            
            # Generate the report first
            filepath = self.generate_shift_report(shift)
            if not filepath:
                logger.error("Failed to generate report for save operation")
                return None
            
            # Create a default filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_name = f"shift_report_{shift.user.username}_{timestamp}.xlsx"
            
            # Open file dialog for save location
            if parent_widget:
                save_path, _ = QFileDialog.getSaveFileName(
                    parent_widget,
                    "Save Shift Report As",
                    default_name,
                    "Excel Files (*.xlsx);;All Files (*)"
                )
            else:
                # Fallback to a default location if no parent widget
                save_path = str(self.reports_dir / default_name)
            
            if save_path:
                # Copy the generated file to the new location
                import shutil
                shutil.copy2(filepath, save_path)
                
                logger.info(f"Report saved to: {save_path}")
                return save_path
            else:
                logger.info("Save operation cancelled by user")
                return None
                
        except Exception as e:
            logger.error(f"Error saving report: {str(e)}")
            return None
    
    def is_excel_available(self) -> bool:
        """
        Check if Excel functionality is available.
        
        Returns:
            bool: True if Excel functionality is available, False otherwise
        """
        # Dynamically check Excel availability instead of relying on static flag
        return check_excel_availability()
    
    def get_excel_status_message(self) -> str:
        """
        Get a status message about Excel functionality availability.
        
        Returns:
            str: Status message about Excel functionality
        """
        if self.is_excel_available():
            return "Excel functionality is available and ready to use."
        else:
            return "Excel functionality is not available. Install openpyxl: pip install openpyxl"
    
    def create_product_import_template(self, filepath: str = None) -> str:
        """
        Create a sample Excel template for product import.
        
        Args:
            filepath: Optional filepath to save the template
            
        Returns:
            str: Path to the created template file
        """
        if not self.is_excel_available():
            logger.error("Excel functionality not available. Install openpyxl: pip install openpyxl")
            return None
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Products"
            
            # Define headers
            headers = ['Name', 'Description', 'Price', 'Category', 'Barcode', 'Image Path']
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Add sample data
            sample_data = [
                ['Sample Product 1', 'This is a sample product description', 9.99, 'Food', '1234567890', ''],
                ['Sample Product 2', 'Another sample product', 15.50, 'Beverage', '0987654321', ''],
                ['Sample Product 3', 'Third sample product', 25.00, 'Dessert', '', '']
            ]
            
            for row, data in enumerate(sample_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = _get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Generate filename if not provided
            if not filepath:
                filepath = "product_import_template.xlsx"
            
            # Save the file
            wb.save(filepath)
            wb.close()
            
            logger.info(f"Product import template created: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error creating product import template: {e}")
            return None
    
    def create_category_import_template(self, filepath: str = None) -> str:
        """
        Create a sample Excel template for category import.
        
        Args:
            filepath: Optional filepath to save the template
            
        Returns:
            str: Path to the created template file
        """
        if not self.is_excel_available():
            logger.error("Excel functionality not available. Install openpyxl: pip install openpyxl")
            return None
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Categories"
            
            # Define headers
            headers = ['Name', 'Description', 'Tax Rate (%)']
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Add sample data
            sample_data = [
                ['Food', 'Food items and meals', 14.0],
                ['Beverage', 'Drinks and beverages', 14.0],
                ['Dessert', 'Desserts and sweets', 14.0],
                ['Other', 'Other miscellaneous items', 14.0]
            ]
            
            for row, data in enumerate(sample_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = _get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Generate filename if not provided
            if not filepath:
                filepath = "category_import_template.xlsx"
            
            # Save the file
            wb.save(filepath)
            wb.close()
            
            logger.info(f"Category import template created: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error creating category import template: {e}")
            return None

    def get_report_preview(self, shift: Shift) -> Dict:
        """
        Get a preview of the report data without generating the file.
        
        Args:
            shift: The shift object containing shift data
            
        Returns:
            Dict: Preview data including summary statistics
        """
        try:
            total_sales = self._get_shift_total_sales(shift)
            sales_data = self._get_shift_sales_data(shift)
            product_data = self._get_shift_product_data(shift)
            
            expected_cash = shift.opening_amount + total_sales
            
            preview = {
                'shift_id': shift.id,
                'cashier_name': shift.user.full_name or shift.user.username,
                        'open_time': shift.open_time.strftime('%Y-%m-%d %I:%M:%S %p'),
        'close_time': shift.close_time.strftime('%Y-%m-%d %I:%M:%S %p') if shift.close_time else 'Not closed',
                'opening_amount': shift.opening_amount,
                'total_sales': total_sales,
                'expected_cash': expected_cash,
                'total_transactions': len(sales_data),
                'total_items': sum(sale['total_items'] for sale in sales_data),
                'average_transaction': total_sales / len(sales_data) if sales_data else 0,
                'products_sold': len(product_data),
                'total_products_quantity': sum(product['quantity'] for product in product_data)
            }
            
            return preview
            
        except Exception as e:
            logger.error(f"Error generating report preview: {str(e)}")
            return {} 